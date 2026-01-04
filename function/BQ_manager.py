from .DB_manager import DB_Manager
import openpyxl

class BQ_Manager:
    def __init__(self, db_manager: DB_Manager):
        self.db = db_manager

    def add_bq_item(self, bq_id, bill, section, page, item, description, qty, unit, rate, discount, trade, remark):
        data = {
            "BQ ID": bq_id,
            "Bill": bill,
            "Section": section,
            "Page": page,
            "Item": item,
            "description": description,
            "Qty": qty,
            "Unit": unit,
            "Rate": rate,
            "Discount": discount,
            "Trade": trade,
            "Remark": remark
        }
        return self.db.insert("Main Contract BQ", data)

    def get_bq_item(self, bq_id):
        query = '''SELECT * FROM "Main Contract BQ" WHERE "BQ ID" = ?'''
        return self.db.fetch_one(query, (bq_id,))

    def get_all_bq_items(self):
        query = '''SELECT * FROM "Main Contract BQ"'''
        return self.db.fetch_all(query)

    def update_bq_item(self, bq_id, data):
        self.db.update("Main Contract BQ", data, '''"BQ ID" = ?''', (bq_id,))

    def delete_bq_item(self, bq_id):
        self.db.delete("Main Contract BQ", '''"BQ ID" = ?''', (bq_id,))

    def get_next_bq_id(self):
        query = '''SELECT "BQ ID" FROM "Main Contract BQ" ORDER BY rowid DESC LIMIT 1'''
        result = self.db.fetch_one(query)
        if result:
            last_id = result['BQ ID']
            num = int(last_id[2:]) + 1  # Assuming format BQ001
        else:
            num = 1
        return f"BQ{num:03d}"

    # Methods for import/export can be added here or in UI

    def import_from_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]  # Assume first row is headers
        
        imported_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            bq_id = data.get("BQ ID")
            if not bq_id:
                continue  # Skip if no BQ ID
            
            # Check if exists
            existing = self.get_bq_item(bq_id)
            if existing:
                # Update
                update_data = {
                    "Bill": data.get("Bill", ""),
                    "Section": data.get("Section", ""),
                    "Page": data.get("Page", ""),
                    "Item": data.get("Item", ""),
                    "description": data.get("Description", ""),
                    "Qty": data.get("Qty", 0),
                    "Unit": data.get("Unit", ""),
                    "Rate": data.get("Rate", 0),
                    "Discount": data.get("Discount", 0),
                    "Trade": data.get("Trade", ""),
                    "Remark": data.get("Remark", "")
                }
                self.update_bq_item(bq_id, update_data)
            else:
                # Insert
                self.add_bq_item(
                    bq_id,
                    data.get("Bill", ""),
                    data.get("Section", ""),
                    data.get("Page", ""),
                    data.get("Item", ""),
                    data.get("Description", ""),
                    data.get("Qty", 0),
                    data.get("Unit", ""),
                    data.get("Rate", 0),
                    data.get("Discount", 0),
                    data.get("Trade", ""),
                    data.get("Remark", "")
                )
            imported_count += 1
        
        return imported_count

    def export_to_excel(self, file_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        headers = ["BQ ID", "Bill", "Section", "Page", "Item", "Description", "Qty", "Unit", "Rate", "Discount", "Trade", "Remark"]
        sheet.append(headers)
        
        items = self.get_all_bq_items()
        for item in items:
            row = [
                item.get("BQ ID", ""),
                item.get("Bill", ""),
                item.get("Section", ""),
                item.get("Page", ""),
                item.get("Item", ""),
                item.get("description", ""),
                item.get("Qty", ""),
                item.get("Unit", ""),
                item.get("Rate", ""),
                item.get("Discount", ""),
                item.get("Trade", ""),
                item.get("Remark", "")
            ]
            sheet.append(row)
        
        wb.save(file_path)
        return len(items)